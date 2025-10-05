from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime
import os
import traceback

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Needed for flash messages
generated_file = "generated/updated_trim.xlsx"


def generate_trim_sheet_152(regn, pilot_weight, pax_weight, fuel_left, fuel_right):
    try:
        wb = load_workbook('master_trim_152.xlsx')
        ws = wb.active

        ws["B1"] = datetime.now().strftime("%d/%m/%Y")
        ws["E2"] = regn

        if regn == "IAU":
            c5, e5 = 1173.96, 31.47
        elif regn == "NNN":
            c5, e5 = 1219, 31.6
        elif regn == "PSS":
            c5, e5 = 1201, 31.09
        else:
            c5, e5 = 0, 0

        ws["C5"] = round(c5, 2)
        ws["E5"] = round(e5, 2)
        ws["G5"] = round(c5 * e5, 2)

        ws["C6"] = round(pilot_weight, 2)
        ws["C7"] = round(pax_weight, 2)

        e6 = ws["E6"].value or 0
        e7 = ws["E7"].value or 0

        ws["G6"] = round(pilot_weight * e6, 2)
        ws["G7"] = round(pax_weight * e7, 2)

        ws["B18"] = round(c5 + pilot_weight + pax_weight, 2)
        ws["G11"] = round(ws["G5"].value + ws["G6"].value + ws["G7"].value, 2)

        ws["B13"] = round(fuel_left, 2)
        ws["B14"] = round(fuel_right, 2)
        ws["B15"] = round(fuel_left + fuel_right, 2)

        ws["C13"] = round(fuel_left * 1.58, 2)
        ws["C14"] = round(fuel_right * 1.58, 2)
        ws["C15"] = round(ws["C13"].value + ws["C14"].value, 2)

        e15 = ws["E15"].value or 0
        ws["G15"] = round(ws["C15"].value * e15, 2)

        ws["G12"] = round(ws["G15"].value, 2)
        ws["G13"] = round(ws["G11"].value + ws["G12"].value, 2)

        ws["B20"] = round(ws["C5"].value + ws["C6"].value + ws["C7"].value + ws["C13"].value + ws["C14"].value, 2)
        ws["E20"] = "Y" if ws["B20"].value < 1670 else "N"

        if ws["B20"].value and ws["G13"].value:
            ws["B21"] = round(ws["G13"].value / ws["B20"].value, 2)

        cg_val = ws["B21"].value
        ws["E21"] = "Y" if cg_val is not None and 31 <= cg_val <= 36.5 else "N"

        os.makedirs("generated", exist_ok=True)
        wb.save(generated_file)

        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=24, max_col=7)]
        return data, None
    except Exception as e:
        return None, f"Error in C152 calculation: {str(e)}"


def generate_trim_sheet_172(regn, pilot_weight, pax_weight, fuel_left, fuel_right):
    try:
        wb = load_workbook('master_trim_172.xlsx')
        ws = wb.active

        ws["B1"] = datetime.now().strftime("%d/%m/%Y")

        if regn == "AGH":
            ws["D2"] = "VT-AGH"
            c5, e5 = 1697, 39.29
        elif regn == "PFA":
            ws["D2"] = "VT-PFA"
            c5, e5 = 1701, 38.79
        else:
            c5, e5 = 0, 0

        ws["C5"] = round(c5, 2)
        ws["E5"] = round(e5, 2)
        ws["G5"] = round(c5 * e5, 2)

        ws["C6"] = round(pilot_weight, 2)
        ws["C7"] = round(pax_weight, 2)

        e6 = ws["E6"].value or 0
        e7 = ws["E7"].value or 0
        ws["G6"] = round(pilot_weight * e6, 2)
        ws["G7"] = round(pax_weight * e7, 2)

        ws["B15"] = round(fuel_left, 2)
        ws["B16"] = round(fuel_right, 2)

        ws["C15"] = round(fuel_left * 1.58, 2)
        ws["C16"] = round(fuel_right * 1.58, 2)

        ws["B17"] = round(fuel_left + fuel_right, 2)
        ws["C17"] = round(ws["C15"].value + ws["C16"].value, 2)

        e17 = ws["E17"].value or 0
        ws["G17"] = round(ws["C17"].value * e17, 2)

        ws["G13"] = round(ws["G5"].value + ws["G6"].value + ws["G7"].value, 2)
        ws["G14"] = round(ws["G17"].value, 2)
        ws["G15"] = round(ws["G13"].value + ws["G14"].value, 2)

        ws["B19"] = round(c5 + pilot_weight + pax_weight, 2)
        ws["B21"] = round(c5 + pilot_weight + pax_weight + ws["C17"].value, 2)

        ws["E21"] = "Y" if ws["B21"].value <= 2550 else "N"

        if ws["B21"].value and ws["G15"].value:
            ws["B22"] = round(ws["G15"].value / ws["B21"].value, 2)

        cg_val = ws["B22"].value
        ws["E22"] = "Y" if cg_val is not None and 35 <= cg_val <= 47.4 else "N"

        os.makedirs("generated", exist_ok=True)
        wb.save(generated_file)

        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=24, max_col=7)]
        return data, None
    except Exception as e:
        return None, f"Error in C172 calculation: {str(e)}"


def generate_trim_sheet_da40ng(regn, pilot_weight, pax_weight, pax1_weight, pax2_weight, baggage_fwd, baggage_aft, fuel_left, fuel_right):
    try:
        wb = load_workbook('master_trim_DA40NG.xlsx')
        ws = wb.active

        # Update date in F1 (append to existing text)
        current_date = datetime.now().strftime("%d/%m/%Y")
        if ws["F1"].value:
            ws["F1"] = f"{ws['F1'].value} {current_date}"
        else:
            ws["F1"] = current_date

        # Update aircraft registration in G2
        if regn == "PM":
            ws["F2"] = "VT-PMA"
        elif regn == "PRH":
            ws["F2"] = "VT-PRH"

        # Update weights - CORRECTED: D11-D16 to E11-E16
        ws["E11"] = round(float(pilot_weight or 0), 2)
        ws["E12"] = round(float(pax_weight or 0), 2)
        ws["E13"] = round(float(pax1_weight or 0), 2)
        ws["E14"] = round(float(pax2_weight or 0), 2)
        ws["E15"] = round(float(baggage_fwd or 0), 2)
        ws["E16"] = round(float(baggage_aft or 0), 2)

        # Update aircraft specific values
        if regn == "PRH":
            ws["E9"] = 2040.44
            ws["F9"] = 95.98
        elif regn == "PM":
            ws["E9"] = 2032.7
            ws["F9"] = 95.47

        # Safe calculation function
        def safe_multiply(a, b):
            a_val = a if a is not None else 0
            b_val = b if b is not None else 0
            return a_val * b_val

        # Calculate G column values safely - CORRECTED: All D11-D16 to E11-E16
        ws["G9"] = round(safe_multiply(ws["E9"].value, ws["F9"].value), 2)
        ws["G11"] = round(safe_multiply(ws["E11"].value, ws["F11"].value), 2)  # CORRECTED: D11 to E11
        ws["G12"] = round(safe_multiply(ws["E12"].value, ws["F12"].value), 2)  # CORRECTED: D12 to E12
        ws["G13"] = round(safe_multiply(ws["E13"].value, ws["F13"].value), 2)  # CORRECTED: D13 to E13
        ws["G14"] = round(safe_multiply(ws["E14"].value, ws["F14"].value), 2)  # CORRECTED: D14 to E14
        ws["G15"] = round(safe_multiply(ws["E15"].value, ws["F15"].value), 2)  # CORRECTED: D15 to E15
        ws["G16"] = round(safe_multiply(ws["E16"].value, ws["F16"].value), 2)  # CORRECTED: D16 to E16

        # Update fuel values
        ws["C17"] = round(float(fuel_left or 0), 2)
        ws["C18"] = round(float(fuel_right or 0), 2)
        
        ws["E17"] = round(float(fuel_left or 0) * 1.58, 2)
        ws["E18"] = round(float(fuel_right or 0) * 1.58, 2)

        # Update total fuel in C19 (append to existing text)
        total_fuel_liters = (float(fuel_left or 0)) + (float(fuel_right or 0))
        if ws["C19"].value:
            ws["C19"] = f"{ws['C19'].value} {total_fuel_liters}"
        else:
            ws["C19"] = str(total_fuel_liters)

        ws["E19"] = round((ws["E9"].value or 0) + (ws["E10"].value or 0) + (ws["E11"].value or 0) + (ws["E12"].value or 0) + (ws["E13"].value or 0) + (ws["E14"].value or 0) + (ws["E15"].value or 0) + (ws["E16"].value or 0) + (ws["E17"].value or 0) + (ws["E18"].value or 0), 2)

        # Calculate G17, G18, G19 safely
        ws["G17"] = round(safe_multiply(ws["E17"].value, ws["F17"].value), 2)
        ws["G18"] = round(safe_multiply(ws["E18"].value, ws["F18"].value), 2)
        
        # Sum all G values from G9 to G18 safely
        g_values = [
            ws["G9"].value, ws["G10"].value, ws["G11"].value, 
            ws["G12"].value, ws["G13"].value, ws["G14"].value,
            ws["G15"].value, ws["G16"].value, ws["G17"].value,
            ws["G18"].value
        ]
        g_sum = sum([val if val is not None else 0 for val in g_values])
        ws["G19"] = round(g_sum, 2)

        # Update B3
        ws["B3"] = round((ws["E17"].value or 0) + (ws["E18"].value or 0), 2)

        # Calculate C24 (CG ratio)
        e19_value = ws["E19"].value or 0
        g19_value = ws["G19"].value or 0
        if e19_value and e19_value != 0:
            C24_value = round(g19_value / e19_value, 2)
            if ws["B24"].value:
                ws["B24"] = f"{ws['B24'].value} {C24_value}"
            else:
                ws["B24"] = str(C24_value)
        else:
            C24_value = 0

        # Calculate C21 (Zero Fuel CG)
        e_values = [
            ws["E9"].value, ws["E10"].value, ws["E11"].value,  # CORRECTED: E11
            ws["E12"].value, ws["E13"].value, ws["E14"].value,  # CORRECTED: E12-E14
            ws["E15"].value, ws["E16"].value  # CORRECTED: E15-E16
        ]
        g_zero_fuel_values = [
            ws["G9"].value, ws["G10"].value, ws["G11"].value,
            ws["G12"].value, ws["G13"].value, ws["G14"].value,
            ws["G15"].value, ws["G16"].value
        ]
        
        e_sum = sum([val if val is not None else 0 for val in e_values])
        g_sum_zero_fuel = sum([val if val is not None else 0 for val in g_zero_fuel_values])
        
        if e_sum and e_sum != 0:
            C21_value = round(g_sum_zero_fuel / e_sum, 2)
            if ws["B21"].value:
                ws["B21"] = f"{ws['B21'].value} {C21_value}"
            else:
                ws["B21"] = str(C21_value)
        else:
            C21_value = 0

        # Calculate C27 (Average CG)
        if C24_value and C21_value:
            C27_value = round((C24_value + C21_value) / 2, 2)
            if ws["B27"].value:
                ws["B27"] = f"{ws['B27'].value} {C27_value}"
            else:
                ws["B27"] = str(C27_value)

        #3update g4
        ws["F4"] = round(sum([ws[f"E{i}"].value or 0 for i in range(9, 17)]), 2)

        os.makedirs("generated", exist_ok=True)
        wb.save(generated_file)

        data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=30, max_col=7)]
        return data, None
    except Exception as e:
        return None, f"Error in DA40NG calculation: {str(e)}\n{traceback.format_exc()}"


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            regn = request.form['regn']
            
            # Get weights with default 0 if empty
            pilot_weight = float(request.form['pilot_weight_lbs'] or 0)
            pax_weight = float(request.form['pax_weight_lbs'] or 0)
            fuel_left = float(request.form['fuel_left'] or 0)
            fuel_right = float(request.form['fuel_right'] or 0)
            
            # DA40NG specific weights - use get() with default
            pax1_weight = float(request.form.get('pax1_weight_lbs', 0) or 0)
            pax2_weight = float(request.form.get('pax2_weight_lbs', 0) or 0)
            baggage_fwd = float(request.form.get('baggage_fwd_lbs', 0) or 0)
            baggage_aft = float(request.form.get('baggage_aft_lbs', 0) or 0)

            data = None
            error = None

            if regn in ["IAU", "NNN", "PSS"]:
                data, error = generate_trim_sheet_152(regn, pilot_weight, pax_weight, fuel_left, fuel_right)
            elif regn in ["AGH", "PFA"]:
                data, error = generate_trim_sheet_172(regn, pilot_weight, pax_weight, fuel_left, fuel_right)
            elif regn in ["PM", "PRH"]:
                data, error = generate_trim_sheet_da40ng(regn, pilot_weight, pax_weight, pax1_weight, pax2_weight, baggage_fwd, baggage_aft, fuel_left, fuel_right)
            else:
                return render_template("error.html", error="Invalid Aircraft Registration"), 400

            if error:
                return render_template("error.html", error=error), 500

            return render_template("table.html", data=data, excel_available=True)

        except Exception as e:
            error_msg = f"Form processing error: {str(e)}\n{traceback.format_exc()}"
            return render_template("error.html", error=error_msg), 500

    return render_template("index.html")


@app.route('/download_excel')
def download_excel():
    try:
        return send_file(generated_file, as_attachment=True)
    except Exception as e:
        return render_template("error.html", error=f"File download error: {str(e)}"), 404


# Add error template route
@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error="Page not found"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error="Internal server error"), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
